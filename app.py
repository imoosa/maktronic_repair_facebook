from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from functools import wraps
import sqlite3, os, uuid, datetime, json, io, hmac, hashlib
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import razorpay
import requests
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

app = Flask(__name__)
app.secret_key = 'maktronics-secret-key-2024'
app.config['UPLOAD_FOLDER'] = os.path.join('static', 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}

# ─── RAZORPAY CONFIG ──────────────────────────────────────────────────────────
RAZORPAY_KEY_ID     = 'rzp_test_SmNSnbXRbFyUas'
RAZORPAY_KEY_SECRET = 'mUiDDcYZ72EzW7pUvtCi4asH'

razorpay_client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))

# ─── WHATSAPP DIRECT (META) CONFIG ──────────────────────────────────────────
WA_TOKEN        = os.environ.get('WA_TOKEN', 'EAAzlNyldxPkBRWA3GNmgrcCLPWZC3LU5Grj7IU8os2nn4U99FkHdkI0PIaCEp4fhjZAr2qktmDQzKdjFbZBs0pOzYJ6kyoQyWe7QjeEGZB1nCYvqUO88wGmh8ON2jrEPHTGz30lvqc6SR3FLpxw4m16qU5FPzREBsEyR3SNvG3s7atoZCb4zb6gH7WVPznNBJDQZDZD')           # Permanent System User token
WA_PHONE_ID     = os.environ.get('WA_PHONE_ID', '1082320074969299')        
WA_VERIFY_TOKEN = os.environ.get('WA_VERIFY_TOKEN', 'maktronics_verify_7253')  # Any string you choose
WA_API_VERSION  = 'v19.0'
WA_API_URL      = f'https://graph.facebook.com/{WA_API_VERSION}/{WA_PHONE_ID}/messages'


# ─── ACCOUNTS DEPARTMENT WHATSAPP NUMBER ─────────────────────────────────────
ACCOUNTS_WHATSAPP_NUMBER = os.environ.get('ACCOUNTS_WHATSAPP_NUMBER', '918551872118')  # ← replace with real number

# ─── PUBLIC BASE URL ──────────────────────────────────────────────────────────
# AiSensy must reach your server from the internet to fetch media (PDF invoices).
#   Production:  export PUBLIC_BASE_URL=https://yourdomain.com
#   Local dev:   export PUBLIC_BASE_URL=https://abc123.ngrok.io  (use ngrok)
PUBLIC_BASE_URL = os.environ.get('PUBLIC_BASE_URL', 'https://maktronic-repair-3.onrender.com').rstrip('/')

def build_public_url(path):
    """Return a publicly accessible URL. path must start with '/'.
    Returns None if PUBLIC_BASE_URL is not set (media attachment skipped)."""
    if not PUBLIC_BASE_URL:
        print(f'[URL] WARNING: PUBLIC_BASE_URL not set — skipping media for: {path}')
        return None
    return f'{PUBLIC_BASE_URL}{path}'


# Template IDs from approved templates
TEMPLATE_IDS = {
    'job_created': 'job_creation',
    'not_repairable': 'not_repairable',
    'estimate_sent': 'estimate_sents',
    'estimate_approved_confirmation': 'estimate_approved',
    'estimate_rejected_confirmation': 'estimate_rejection',
    'repair_completed_invoice': 'dev_invoice_sent',              # Razorpay payment link invoice
    'repair_completed_invoice_not_razorpay': 'dev_invoice_sent_other',  # Non-Razorpay invoice
    'payment_received': 'payment_received',
    'product_dispatched': 'product_dispatch',
    'accounts_department': 'accounts_department',  # Notify accounts to generate invoice
}

# Helper function to format phone number
def format_phone_number(phone):
    if not phone:
        return None
    phone = str(phone).strip()
    phone = ''.join(filter(str.isdigit, phone))
    if phone.startswith('0'):
        phone = phone[1:]
    if not phone.startswith('91') and len(phone) == 10:
        phone = '91' + phone
    return phone

def send_whatsapp_template(to_number, template_name, variables, media_url=None):
    """Send a WhatsApp template message via Meta Graph API directly."""
    if not WA_TOKEN or not WA_PHONE_ID:
        return {'success': False, 'error': 'WA_TOKEN or WA_PHONE_ID not configured'}

    to_number = format_phone_number(to_number)  # reuse your existing helper
    if not to_number:
        return {'success': False, 'error': 'Invalid phone number'}

    # Build template components (text parameters)
    text_params = [{"type": "text", "text": str(v)} for v in variables]

    components = [{"type": "body", "parameters": text_params}]

    # If there's a media attachment (e.g. PDF invoice), add header component
    if media_url:
        components.insert(0, {
            "type": "header",
            "parameters": [{
                "type": "document",
                "document": {
                    "link": media_url,
                    "filename": f"invoice_{datetime.datetime.now().strftime('%Y%m%d')}.pdf"
                }
            }]
        })

    payload = {
        "messaging_product": "whatsapp",
        "to": to_number,
        "type": "template",
        "template": {
            "name": template_name,   # exact name from Meta template manager
            "language": {"code": "en"},
            "components": components
        }
    }

    headers = {
        "Authorization": f"Bearer {WA_TOKEN}",
        "Content-Type": "application/json"
    }

    try:
        print(f"[WhatsApp] Sending template '{template_name}' to {to_number}")
        response = requests.post(WA_API_URL, json=payload, headers=headers, timeout=30)
        response.raise_for_status()
        result = response.json()
        msg_id = result.get('messages', [{}])[0].get('id')
        print(f"[WhatsApp] Success. Message ID: {msg_id}")
        return {'success': True, 'response': result, 'message_id': msg_id}
    except requests.exceptions.HTTPError as e:
        err_body = e.response.text if e.response else str(e)
        print(f"[WhatsApp Error] {e} | {err_body}")
        return {'success': False, 'error': str(e), 'details': err_body}
    except Exception as e:
        print(f"[WhatsApp Error] {str(e)}")
        return {'success': False, 'error': str(e)}

def send_accounts_department_notification(job, total_amount, parts_cost, labour_cost, tech_name='Technician'):
    """Notify accounts department to generate invoice after tech marks repair done.
      Template: accounts_department
      {{1}} customer_name  {{2}} job_id  {{3}} item_type
      {{4}} ₹total  {{5}} ₹parts  {{6}} ₹labour  {{7}} technician_name
    """
    if not ACCOUNTS_WHATSAPP_NUMBER:
        print('[WhatsApp] ACCOUNTS_WHATSAPP_NUMBER not set — skipping accounts notification')
        return {'success': False, 'error': 'ACCOUNTS_WHATSAPP_NUMBER not configured'}

    variables = [
        job.get('customer_name', 'Customer'),   
        job.get('job_id', 'N/A'),               
        job.get('item_type', 'Device'),         
        f"\u20b9{total_amount:.2f}",           
        f"\u20b9{parts_cost:.2f}",             
        f"\u20b9{labour_cost:.2f}",            
        tech_name,                              
    ]

    # Build payload manually — send to accounts number, not customer
    url = "https://backend.aisensy.com/campaign/t1/api/v2"
    payload = {
        "apiKey": AISENSY_API_KEY,
        "campaignName": TEMPLATE_IDS['accounts_department'],
        "destination": ACCOUNTS_WHATSAPP_NUMBER,
        "userName": variables[0],
        "source": "api",
        "templateParams": variables,
        "tags": [],
        "attributes": {}
    }
    headers = {'Content-Type': 'application/json'}
    try:
        print(f"[WhatsApp Accounts] Sending invoice request for job {job.get('job_id')} to {ACCOUNTS_WHATSAPP_NUMBER}")
        print(f"[WhatsApp Accounts] Payload: {payload}")
        response = requests.post(url, json=payload, headers=headers, timeout=30)
        print(f"[WhatsApp Accounts] Status: {response.status_code} | Body: {response.text}")
        response.raise_for_status()
        print(f"[WhatsApp Accounts] Success: {response.json()}")
        return {'success': True, 'response': response.json()}
    except requests.exceptions.HTTPError as e:
        err_body = e.response.text if e.response else str(e)
        print(f"[WhatsApp Accounts Error] HTTP {e.response.status_code if e.response else '?'} | Body: {err_body}")
        return {'success': False, 'error': str(e), 'details': err_body}
    except Exception as e:
        print(f"[WhatsApp Accounts Error] {str(e)}")
        return {'success': False, 'error': str(e)}

def send_job_created_notification(job):
    """Template 1: Job Created & Barcode Generated"""
    variables = [
        job.get('customer_name', 'Customer'),
        job.get('job_id', 'N/A'),
        job.get('item_type', 'Device'),
        job.get('item_description', 'Issue description') or 'General Service'
    ]
    return send_whatsapp_template(job.get('customer_phone'), TEMPLATE_IDS['job_created'], variables)

def send_not_repairable_notification(job, reason):
    """Template 2: Device Not Repairable"""
    variables = [
        job.get('customer_name', 'Customer'),
        job.get('job_id', 'N/A'),
        job.get('item_type', 'Device'),
        reason
    ]
    return send_whatsapp_template(job.get('customer_phone'), TEMPLATE_IDS['not_repairable'], variables)

def send_estimate_notification(job, estimate_amount, parts_cost, labour_cost):
    """Template 3: Estimate Sent (with YES/NO approval flow)"""
    variables = [
        job.get('customer_name', 'Customer'),
        job.get('job_id', 'N/A'),
        job.get('item_type', 'Device'),
        f"{estimate_amount:.2f}",
        f"{parts_cost:.2f}",
        f"{labour_cost:.2f}"
    ]
    return send_whatsapp_template(job.get('customer_phone'), TEMPLATE_IDS['estimate_sent'], variables)

def send_estimate_approved_confirmation(job):
    """Template 4: Estimate Approved Confirmation"""
    variables = [
        job.get('customer_name', 'Customer'),
        job.get('job_id', 'N/A')
    ]
    return send_whatsapp_template(job.get('customer_phone'), TEMPLATE_IDS['estimate_approved_confirmation'], variables)

def send_estimate_rejected_confirmation(job):
    """Template 5: Estimate Rejected Confirmation"""
    variables = [
        job.get('customer_name', 'Customer'),
        job.get('job_id', 'N/A')
    ]
    return send_whatsapp_template(job.get('customer_phone'), TEMPLATE_IDS['estimate_rejected_confirmation'], variables)

def send_invoice_ready_notification(job, total_amount, parts_cost, labour_cost, invoice_url=None):
    """Template dev_invoice_sent — 8 params matching AiSensy template:
      {{1}} name  {{2}} job_id  {{3}} item_type
      {{4}} ₹total  {{5}} ₹parts  {{6}} ₹labour
      {{7}} job_id (for payment URL)  {{8}} payment_token (for payment URL)
    """
    variables = [
        job.get('customer_name', 'Customer'),   # {{1}}
        job.get('job_id', 'N/A'),               # {{2}}
        job.get('item_type', 'Device'),         # {{3}}
        f"\u20b9{total_amount:.2f}",           # {{4}}
        f"\u20b9{parts_cost:.2f}",             # {{5}}
        f"\u20b9{labour_cost:.2f}",            # {{6}}
        job.get('job_id', 'N/A'),               # {{7}}
        job.get('payment_token', ''),           # {{8}}
    ]
    return send_whatsapp_template(job.get('customer_phone'), TEMPLATE_IDS['repair_completed_invoice'], variables, media_url=invoice_url)

def send_invoice_ready_notification_non_razorpay(job, total_amount, parts_cost, labour_cost, payment_method_label, invoice_url=None):
    """Template dev_invoice_sent_other — for Cash, Cheque, Pay Later, Other payment methods.
      No payment link. Params:
      {{1}} name  {{2}} job_id  {{3}} item_type
      {{4}} total  {{5}} parts  {{6}} labour  {{7}} payment_method_label
    """
    variables = [
        job.get('customer_name', 'Customer'),   # {{1}}
        job.get('job_id', 'N/A'),               # {{2}}
        job.get('item_type', 'Device'),         # {{3}}
        f"\u20b9{total_amount:.2f}",           # {{4}}
        f"\u20b9{parts_cost:.2f}",             # {{5}}
        f"\u20b9{labour_cost:.2f}",            # {{6}}
        payment_method_label,                   # {{7}}
    ]
    return send_whatsapp_template(
        job.get('customer_phone'),
        TEMPLATE_IDS['repair_completed_invoice_not_razorpay'],
        variables,
        media_url=invoice_url
    )

def send_payment_received_confirmation(job, amount, transaction_id):
    """Template 7: Payment Received Confirmation"""
    variables = [
        job.get('customer_name', 'Customer'),   # {{1}}
        f"{amount:.2f}",                        # {{2}}
        job.get('job_id', 'N/A'),               # {{3}}
        transaction_id,                          # {{4}}
        datetime.datetime.now().strftime('%d/%m/%Y')  # {{5}}
    ]
    return send_whatsapp_template(
        job.get('customer_phone'),
        TEMPLATE_IDS['payment_received'],
        variables
    )

def send_dispatched_notification(job, courier_name, tracking_number, expected_delivery, media_url=None):
    variables = [
        job.get('customer_name', 'Customer'),
        job.get('job_id', 'N/A'),
        job.get('item_type', 'Device'),
        courier_name or 'Our Courier Partner',
        tracking_number or 'N/A',
        expected_delivery or '3-5 business days'
    ]
    return send_whatsapp_template(
        job.get('customer_phone'),
        TEMPLATE_IDS['product_dispatched'],
        variables,
        media_url=media_url  # ✅ pass receipt image
    )

def send_payment_link_via_whatsapp(job, payment_link, amount):
    """Payment link is embedded in dev_invoice_sent template via {{7}} job_id + {{8}} token.
    This is called after send_invoice_ready_notification so the token is already saved.
    No separate message needed — just logs the link for reference.
    """
    if not payment_link:
        print(f"[WhatsApp] pay_link is None for {job.get('job_id')} — set PUBLIC_BASE_URL env var")
        return {'success': False, 'error': 'payment_link is None (PUBLIC_BASE_URL not set?)'}
    print(f"[WhatsApp] Payment link for {job.get('job_id')}: {payment_link} (embedded in invoice template)")
    return {'success': True, 'note': 'payment link already sent via invoice template'}

@app.context_processor
def inject_globals():
    def _has_perm(perm):
        if not session.get('user_id'):
            return False
        if session.get('role') == 'admin':
            return True
        db = get_db()
        user = db.execute("SELECT permissions FROM users WHERE id=?", (session['user_id'],)).fetchone()
        if not user:
            return False
        try:
            return json.loads(user['permissions'] or '{}').get(perm, False)
        except Exception:
            return False

    def _unread_count():
        if not session.get('user_id'):
            return 0
        db = get_db()
        row = db.execute(
            "SELECT COUNT(*) FROM notifications WHERE recipient_id=? AND is_read=0",
            (session['user_id'],)
        ).fetchone()
        return row[0] if row else 0

    return dict(has_perm=_has_perm, unread_notif_count=_unread_count)

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

DB_PATH = 'maktronics.db'

# ─── STATUS FLOW ──────────────────────────────────────────────────────────────
STATUS_FLOW = {
    'sent_for_inspection': {'label': 'Sent for Inspection', 'icon': '🔍', 'color': '#FFB74D'},
    'not_repairable':      {'label': 'Not Repairable',      'icon': '❌', 'color': '#EF5350'},
    'estimate_sent':       {'label': 'Estimate Sent',       'icon': '📄', 'color': '#FFD54F'},
    'estimate_approved':   {'label': 'Estimate Approved',   'icon': '👍', 'color': '#4DB6AC'},
    'estimate_rejected':   {'label': 'Estimate Rejected',   'icon': '🚫', 'color': '#EF5350'},
    'sent_for_repair':     {'label': 'Sent for Repair',     'icon': '🔧', 'color': '#F06292'},
    'repair_done':         {'label': 'Repair Done',         'icon': '✅', 'color': '#4DB6AC'},
    'invoice_uploaded':    {'label': 'Invoice Uploaded',    'icon': '📎', 'color': '#AED581'},
    'payment_received':    {'label': 'Payment Received',    'icon': '💰', 'color': '#66BB6A'},
    'dispatched':          {'label': 'Dispatched',          'icon': '🚚', 'color': '#4DD0E1'},
    'closed':              {'label': 'Closed',              'icon': '🔒', 'color': '#90A4AE'},
}

STATUS_ORDER = [
    'sent_for_inspection', 'estimate_sent', 'estimate_approved',
    'sent_for_repair', 'repair_done', 'invoice_uploaded',
    'payment_received', 'dispatched', 'closed',
]

# ─── DB SETUP ─────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as db:
        db.executescript('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                role TEXT NOT NULL CHECK(role IN ('admin','technician','manager')),
                name TEXT NOT NULL,
                permissions TEXT DEFAULT "{}",
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS jobs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_id TEXT UNIQUE NOT NULL,
                customer_name TEXT,
                customer_phone TEXT,
                customer_email TEXT,
                item_description TEXT,
                item_type TEXT,
                barcode TEXT,
                status TEXT DEFAULT 'sent_for_inspection',
                assigned_tech_id INTEGER,
                received_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                repair_findings TEXT,
                parts_cost REAL DEFAULT 0,
                labour_cost REAL DEFAULT 0,
                total_amount REAL DEFAULT 0,
                invoice_number TEXT,
                payment_status TEXT DEFAULT 'pending',
                payment_received_at TIMESTAMP,
                tracking_number TEXT,
                dispatch_date TEXT,
                expected_delivery TEXT,
                delivered_at TIMESTAMP,
                notes TEXT,
                invoice_path TEXT,
                courier_name TEXT,
                courier_receipt_path TEXT,
                estimate_amount REAL DEFAULT 0,
                estimate_notes TEXT,
                estimate_sent_at TIMESTAMP,
                estimate_approved_at TIMESTAMP,
                not_repairable_reason TEXT,
                sent_back_to_customer_at TIMESTAMP,
                inspection_findings TEXT,
                invoice_generate_date TIMESTAMP,
                invoice_total_amount REAL DEFAULT 0,
                razorpay_order_id TEXT,
                razorpay_payment_id TEXT,
                payment_link TEXT,
                payment_token TEXT,
                whatsapp_message_id TEXT,
                payment_method TEXT DEFAULT 'razorpay',
                FOREIGN KEY(assigned_tech_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS job_photos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_id TEXT NOT NULL,
                photo_path TEXT NOT NULL,
                uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS job_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_id TEXT NOT NULL,
                action TEXT NOT NULL,
                performed_by INTEGER,
                details TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(performed_by) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS notifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                recipient_id INTEGER NOT NULL,
                job_id TEXT NOT NULL,
                message TEXT NOT NULL,
                is_read INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(recipient_id) REFERENCES users(id)
            );
        ''')

        # Seed default users
        existing = db.execute("SELECT id FROM users WHERE username='admin'").fetchone()
        if not existing:
            db.execute("INSERT INTO users (username, password, role, name) VALUES (?,?,?,?)",
                ('admin', generate_password_hash('admin123'), 'admin', 'Admin User'))
            db.execute("INSERT INTO users (username, password, role, name) VALUES (?,?,?,?)",
                ('tech1', generate_password_hash('tech123'), 'technician', 'Ravi Kumar'))
            db.execute("INSERT INTO users (username, password, role, name) VALUES (?,?,?,?)",
                ('tech2', generate_password_hash('tech456'), 'technician', 'Suresh Patil'))
        db.commit()

# ─── HELPERS ──────────────────────────────────────────────────────────────────
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def generate_job_id():
    now = datetime.datetime.now()
    return f"MAK-{now.strftime('%Y%m%d')}-{str(uuid.uuid4())[:6].upper()}"

def log_action(db, job_id, action, user_id, details=''):
    db.execute("INSERT INTO job_logs (job_id, action, performed_by, details) VALUES (?,?,?,?)",
               (job_id, action, user_id, details))

def send_notification(db, recipient_id, job_id, message):
    db.execute(
        "INSERT INTO notifications (recipient_id, job_id, message) VALUES (?,?,?)",
        (recipient_id, job_id, message)
    )

def notify_all_admins(db, job_id, message):
    admins = db.execute(
        "SELECT id FROM users WHERE role IN ('admin','manager')"
    ).fetchall()
    for admin in admins:
        send_notification(db, admin['id'], job_id, message)

def get_assigned_tech_id(db, job_id):
    row = db.execute("SELECT assigned_tech_id FROM jobs WHERE job_id=?", (job_id,)).fetchone()
    return row['assigned_tech_id'] if row else None

# ─── AUTH ─────────────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') not in ('admin', 'manager'):
            flash('Admin access required.', 'error')
            return redirect(url_for('tech_dashboard'))
        return f(*args, **kwargs)
    return decorated

# ─── WHATSAPP WEBHOOK (Handles YES/NO replies) ───────────────────────────────
# ─── WEBHOOK VERIFICATION (GET) ──────────────────────────────────────────────
@app.route('/webhook/whatsapp', methods=['GET'])
def whatsapp_webhook_verify():
    mode      = request.args.get('hub.mode')
    token     = request.args.get('hub.verify_token')
    challenge = request.args.get('hub.challenge')

    if mode == 'subscribe' and token == WA_VERIFY_TOKEN:
        print('[WhatsApp Webhook] Verified!')
        return challenge, 200
    return 'Forbidden', 403

# ─── WEBHOOK RECEIVER (POST) ─────────────────────────────────────────────────
@app.route('/webhook/whatsapp', methods=['POST'])
def whatsapp_webhook_receive():
    data = request.get_json()
    print(f'[WhatsApp Webhook] Incoming: {json.dumps(data, indent=2)}')

    try:
        entry   = data['entry'][0]
        changes = entry['changes'][0]['value']

        # Incoming message from customer
        if 'messages' in changes:
            msg    = changes['messages'][0]
            from_  = msg['from']        # customer's phone number
            msg_id = msg['id']
            body   = msg.get('text', {}).get('body', '')
            print(f'[WhatsApp] Message from {from_}: {body}')
            # TODO: handle customer replies here (e.g., estimate approval)

        # Message status updates (sent/delivered/read/failed)
        if 'statuses' in changes:
            status = changes['statuses'][0]
            print(f'[WhatsApp] Status update: {status["status"]} for msg {status["id"]}')

    except (KeyError, IndexError) as e:
        print(f'[WhatsApp Webhook] Parse error: {e}')

    return 'OK', 200

# ─── ROUTES ───────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    if 'user_id' in session:
        if session['role'] == 'admin':
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('tech_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            session['name'] = user['name']
            if user['role'] == 'admin':
                return redirect(url_for('admin_dashboard'))
            return redirect(url_for('tech_dashboard'))
        flash('Invalid credentials', 'error')
    return render_template('shared/login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ─── NOTIFICATIONS ────────────────────────────────────────────────────────────
@app.route('/notifications')
@login_required
def notifications():
    db = get_db()
    user_id = session['user_id']
    notifs = db.execute(
        "SELECT * FROM notifications WHERE recipient_id=? ORDER BY created_at DESC LIMIT 50",
        (user_id,)
    ).fetchall()
    db.execute("UPDATE notifications SET is_read=1 WHERE recipient_id=?", (user_id,))
    db.commit()
    return render_template('shared/notifications.html', notifications=notifs)

@app.route('/notifications/mark_read_all', methods=['POST'])
@login_required
def mark_all_notifications_read():
    db = get_db()
    db.execute("UPDATE notifications SET is_read=1 WHERE recipient_id=?", (session['user_id'],))
    db.commit()
    return jsonify({'ok': True})

@app.route('/notifications/mark_read', methods=['POST'])
@login_required
def mark_notification_read():
    notif_id = request.form.get('notif_id')
    db = get_db()
    db.execute(
        "UPDATE notifications SET is_read=1 WHERE id=? AND recipient_id=?",
        (notif_id, session['user_id'])
    )
    db.commit()
    return jsonify({'ok': True})

@app.route('/api/notifications/unread_count')
@login_required
def api_unread_count():
    db = get_db()
    row = db.execute(
        "SELECT COUNT(*) FROM notifications WHERE recipient_id=? AND is_read=0",
        (session['user_id'],)
    ).fetchone()
    return jsonify({'count': row[0] if row else 0})

@app.route('/api/notifications/poll')
@login_required
def api_notifications_poll():
    db = get_db()
    user_id = session['user_id']
    notifs = db.execute(
        "SELECT * FROM notifications WHERE recipient_id=? ORDER BY created_at DESC LIMIT 50",
        (user_id,)
    ).fetchall()
    unread = db.execute(
        "SELECT COUNT(*) FROM notifications WHERE recipient_id=? AND is_read=0",
        (user_id,)
    ).fetchone()[0]
    return jsonify({
        'unread_count': unread,
        'notifications': [
            {
                'id': n['id'],
                'job_id': n['job_id'],
                'message': n['message'],
                'is_read': bool(n['is_read']),
                'created_at': n['created_at'],
            }
            for n in notifs
        ]
    })

@app.route('/api/job/<job_id>/status')
@login_required
def api_job_status(job_id):
    db = get_db()
    job = db.execute(
        "SELECT j.*, u.name as tech_name FROM jobs j LEFT JOIN users u ON j.assigned_tech_id=u.id WHERE j.job_id=?",
        (job_id,)
    ).fetchone()
    if not job:
        return jsonify({'error': 'Job not found'}), 404
    if session.get('role') not in ('admin', 'manager'):
        if job['assigned_tech_id'] != session['user_id']:
            return jsonify({'error': 'Access denied'}), 403
    logs = db.execute(
        """SELECT l.*, u.name as user_name FROM job_logs l
           LEFT JOIN users u ON l.performed_by=u.id
           WHERE l.job_id=? ORDER BY l.created_at DESC LIMIT 20""",
        (job_id,)
    ).fetchall()
    return jsonify({
        'status': job['status'],
        'updated_at': job['updated_at'],
        'tech_name': job['tech_name'],
        'payment_status': job['payment_status'],
        'parts_cost': job['parts_cost'],
        'labour_cost': job['labour_cost'],
        'total_amount': job['total_amount'],
        'invoice_number': job['invoice_number'],
        'payment_link': job['payment_link'],
        'tracking_number': job['tracking_number'],
        'courier_name': job['courier_name'],
        'dispatch_date': job['dispatch_date'],
        'expected_delivery': job['expected_delivery'],
        'repair_findings': job['repair_findings'],
        'inspection_findings': job['inspection_findings'],
        'estimate_amount': job['estimate_amount'],
        'estimate_notes': job['estimate_notes'],
        'not_repairable_reason': job['not_repairable_reason'],
        'logs': [
            {
                'user_name': l['user_name'] or 'System',
                'action': l['action'],
                'details': l['details'],
                'created_at': l['created_at'],
            }
            for l in logs
        ],
    })

# ─── ADMIN ROUTES ─────────────────────────────────────────────────────────────
@app.route('/admin')
@admin_required
def admin_dashboard():
    db = get_db()
    search = request.args.get('search', '').strip()

    stats = {
        'total':           db.execute("SELECT COUNT(*) FROM jobs").fetchone()[0],
        'sent_for_repair': db.execute("SELECT COUNT(*) FROM jobs WHERE status='sent_for_repair'").fetchone()[0],
        'payment_pending': db.execute("SELECT COUNT(*) FROM jobs WHERE payment_status='pending' AND status='invoice_uploaded'").fetchone()[0],
        'dispatched':      db.execute("SELECT COUNT(*) FROM jobs WHERE status='dispatched'").fetchone()[0],
        'closed':          db.execute("SELECT COUNT(*) FROM jobs WHERE status='closed'").fetchone()[0],
    }

    rev_row = db.execute("""
        SELECT
            COALESCE(SUM(total_amount), 0) AS total_revenue,
            COALESCE(SUM(parts_cost), 0) AS total_parts,
            COALESCE(SUM(labour_cost), 0) AS total_labour,
            COALESCE(SUM(CASE WHEN payment_status='paid' THEN total_amount ELSE 0 END), 0) AS paid_revenue
        FROM jobs
    """).fetchone()

    status_rows = db.execute("SELECT status, COUNT(*) as cnt FROM jobs GROUP BY status").fetchall()
    status_counts = {r['status']: r['cnt'] for r in status_rows}

    tech_rows = db.execute("""
        SELECT
            COALESCE(u.name, 'Unassigned') AS name,
            COALESCE(SUM(j.total_amount), 0) AS revenue,
            COALESCE(SUM(j.parts_cost), 0) AS parts,
            COALESCE(SUM(j.labour_cost), 0) AS labour
        FROM jobs j
        LEFT JOIN users u ON j.assigned_tech_id = u.id
        GROUP BY j.assigned_tech_id
        ORDER BY revenue DESC
    """).fetchall()
    tech_revenue = [dict(r) for r in tech_rows]

    monthly_rows = db.execute("""
        SELECT
            strftime('%Y-%m', received_at) AS month,
            COUNT(*) AS jobs,
            COALESCE(SUM(total_amount), 0) AS revenue
        FROM jobs
        WHERE received_at >= date('now', '-6 months')
        GROUP BY month
        ORDER BY month
    """).fetchall()
    monthly = [dict(r) for r in monthly_rows]

    item_rows = db.execute("""
        SELECT item_type, COUNT(*) AS count
        FROM jobs
        WHERE item_type IS NOT NULL AND item_type != ''
        GROUP BY item_type
        ORDER BY count DESC
    """).fetchall()
    item_types = [dict(r) for r in item_rows]

    analytics = {
        'total_revenue': rev_row['total_revenue'],
        'total_parts': rev_row['total_parts'],
        'total_labour': rev_row['total_labour'],
        'paid_revenue': rev_row['paid_revenue'],
        'status_counts': status_counts,
        'tech_revenue': tech_revenue,
        'monthly': monthly,
        'item_types': item_types,
    }

    if search:
        recent_jobs = db.execute("""
            SELECT j.*, u.name as tech_name FROM jobs j
            LEFT JOIN users u ON j.assigned_tech_id = u.id
            WHERE j.job_id LIKE ? OR j.barcode LIKE ? OR j.customer_name LIKE ? OR j.customer_phone LIKE ?
            ORDER BY j.received_at DESC LIMIT 50
        """, (f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%')).fetchall()
    else:
        recent_jobs = db.execute("""
            SELECT j.*, u.name as tech_name FROM jobs j
            LEFT JOIN users u ON j.assigned_tech_id = u.id
            ORDER BY j.received_at DESC LIMIT 20
        """).fetchall()

    technicians = db.execute("SELECT * FROM users WHERE role='technician'").fetchall()
    return render_template('admin/dashboard.html', stats=stats, jobs=recent_jobs,
                           technicians=technicians, status_flow=STATUS_FLOW,
                           search=search, analytics=analytics)

@app.route('/admin/jobs')
@admin_required
def admin_jobs():
    db = get_db()
    status_filter = request.args.get('status', '')
    search = request.args.get('search', '')
    query = """
        SELECT j.*, u.name as tech_name FROM jobs j
        LEFT JOIN users u ON j.assigned_tech_id = u.id
        WHERE 1=1
    """
    params = []
    if status_filter:
        query += " AND j.status=?"
        params.append(status_filter)
    if search:
        query += " AND (j.job_id LIKE ? OR j.barcode LIKE ? OR j.customer_name LIKE ? OR j.customer_phone LIKE ?)"
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%'])
    query += " ORDER BY j.received_at DESC"
    jobs = db.execute(query, params).fetchall()
    technicians = db.execute("SELECT * FROM users WHERE role='technician'").fetchall()
    return render_template('admin/jobs.html', jobs=jobs, status_flow=STATUS_FLOW,
                           technicians=technicians, status_filter=status_filter, search=search)

@app.route('/admin/jobs/export')
@admin_required
def admin_jobs_export():
    """Export all jobs to Excel (.xlsx)"""
    db = get_db()
    status_filter = request.args.get('status', '')
    search = request.args.get('search', '')

    query = """
        SELECT j.job_id, j.customer_name, j.customer_phone, j.customer_email,
               j.item_type, j.item_description, j.status, j.payment_status,
               u.name as tech_name,
               j.parts_cost, j.labour_cost, j.total_amount, j.invoice_total_amount,
               j.invoice_number, j.payment_method,
               j.received_at, j.updated_at, j.dispatch_date, j.tracking_number,
               j.courier_name, j.repair_findings, j.notes
        FROM jobs j
        LEFT JOIN users u ON j.assigned_tech_id = u.id
        WHERE 1=1
    """
    params = []
    if status_filter:
        query += " AND j.status=?"
        params.append(status_filter)
    if search:
        query += " AND (j.job_id LIKE ? OR j.barcode LIKE ? OR j.customer_name LIKE ? OR j.customer_phone LIKE ?)"
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%'])
    query += " ORDER BY j.received_at DESC"
    jobs = db.execute(query, params).fetchall()

    if not OPENPYXL_AVAILABLE:
        # Fallback: CSV export
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        headers = ['Job ID','Customer','Phone','Email','Item Type','Description','Status',
                   'Payment Status','Technician','Parts Cost','Labour Cost','Total Amount',
                   'Invoice Amount','Invoice #','Payment Method','Received At','Updated At',
                   'Dispatch Date','Tracking #','Courier','Repair Findings','Notes']
        writer.writerow(headers)
        for j in jobs:
            writer.writerow([
                j['job_id'], j['customer_name'], j['customer_phone'], j['customer_email'],
                j['item_type'], j['item_description'],
                STATUS_FLOW.get(j['status'], {}).get('label', j['status']),
                j['payment_status'], j['tech_name'],
                j['parts_cost'] or 0, j['labour_cost'] or 0, j['total_amount'] or 0,
                j['invoice_total_amount'] or 0, j['invoice_number'], j['payment_method'] or '',
                j['received_at'], j['updated_at'], j['dispatch_date'],
                j['tracking_number'], j['courier_name'], j['repair_findings'], j['notes']
            ])
        output.seek(0)
        filename = f"jobs_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return send_file(io.BytesIO(output.getvalue().encode()), mimetype='text/csv',
                         as_attachment=True, download_name=filename)

    # ── Build Excel workbook ───────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jobs Export"

    # Color palette
    HEADER_FILL   = PatternFill("solid", fgColor="1A1D2E")
    ALT_FILL      = PatternFill("solid", fgColor="F4F6FF")
    ACCENT_COLOR  = "4A9EFF"
    HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT    = Font(bold=True, color="1A1D2E", size=14)
    MONEY_FONT    = Font(bold=True, color="22C55E", size=11)
    thin_side     = Side(style='thin', color='D1D5DB')
    thin_border   = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells('A1:V1')
    title_cell = ws['A1']
    title_cell.value = f"Maktronics — Jobs Export  ({datetime.datetime.now().strftime('%d %b %Y, %H:%M')})"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill("solid", fgColor="EEF4FF")
    ws.row_dimensions[1].height = 30

    # ── Summary row ───────────────────────────────────────────────────────────
    ws.merge_cells('A2:V2')
    total_amt = sum((j['total_amount'] or 0) for j in jobs)
    paid_count = sum(1 for j in jobs if j['payment_status'] == 'paid')
    summary = ws['A2']
    summary.value = (f"Total Jobs: {len(jobs)}   |   "
                     f"Paid: {paid_count}   |   "
                     f"Pending: {len(jobs)-paid_count}   |   "
                     f"Total Revenue: ₹{total_amt:,.2f}")
    summary.font = Font(italic=True, color="6B7280", size=10)
    summary.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    # ── Column headers ────────────────────────────────────────────────────────
    COLUMNS = [
        ('Job ID',           16), ('Customer',        20), ('Phone',          14),
        ('Email',            24), ('Item Type',        14), ('Description',    28),
        ('Status',           18), ('Payment Status',  16), ('Technician',      18),
        ('Parts (₹)',        12), ('Labour (₹)',       12), ('Total (₹)',       13),
        ('Invoice Amt (₹)',  14), ('Invoice #',        16), ('Payment Method',  18),
        ('Received At',      18), ('Updated At',       18), ('Dispatch Date',  14),
        ('Tracking #',       18), ('Courier',          16), ('Repair Findings', 30),
        ('Notes',            30),
    ]
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[3].height = 28

    # ── Data rows ─────────────────────────────────────────────────────────────
    money_fmt = '#,##0.00'
    for row_idx, j in enumerate(jobs, start=4):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        status_label = STATUS_FLOW.get(j['status'], {}).get('label', j['status'] or '')
        row_data = [
            j['job_id'], j['customer_name'] or '', j['customer_phone'] or '',
            j['customer_email'] or '', j['item_type'] or '', j['item_description'] or '',
            status_label, (j['payment_status'] or 'pending').replace('_', ' ').title(),
            j['tech_name'] or 'Unassigned',
            j['parts_cost'] or 0, j['labour_cost'] or 0, j['total_amount'] or 0,
            j['invoice_total_amount'] or 0, j['invoice_number'] or '',
            (j['payment_method'] or '').replace('_', ' ').title(),
            j['received_at'] or '', j['updated_at'] or '', j['dispatch_date'] or '',
            j['tracking_number'] or '', j['courier_name'] or '',
            j['repair_findings'] or '', j['notes'] or '',
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if fill:
                cell.fill = fill
            # Money columns: 10, 11, 12, 13
            if col_idx in (10, 11, 12, 13) and isinstance(value, (int, float)):
                cell.number_format = money_fmt
                cell.font = Font(color="22C55E", bold=(col_idx == 12))

        ws.row_dimensions[row_idx].height = 20

    # Freeze header rows
    ws.freeze_panes = 'A4'

    # Auto-filter on header row
    ws.auto_filter.ref = f"A3:V{max(3, len(jobs)+3)}"

    # ── Save and send ──────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"maktronics_jobs_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@app.route('/admin/jobs/new', methods=['GET', 'POST'])
@admin_required
def admin_new_job():
    if request.method == 'POST':
        db = get_db()
        job_id = generate_job_id()
        barcode = job_id
        tech_id = request.form.get('assigned_tech_id') or None

        data = {
            'job_id': job_id,
            'customer_name': request.form.get('customer_name'),
            'customer_phone': request.form.get('customer_phone'),
            'customer_email': request.form.get('customer_email'),
            'item_description': request.form.get('item_description'),
            'item_type': request.form.get('item_type'),
            'barcode': barcode,
            'assigned_tech_id': tech_id,
            'status': 'sent_for_inspection',
            'notes': request.form.get('notes'),
        }
        db.execute("""
            INSERT INTO jobs (job_id, customer_name, customer_phone, customer_email,
            item_description, item_type, barcode, assigned_tech_id, status, notes)
            VALUES (:job_id,:customer_name,:customer_phone,:customer_email,
            :item_description,:item_type,:barcode,:assigned_tech_id,:status,:notes)
        """, data)

        photos = request.files.getlist('photos')
        for photo in photos[:3]:
            if photo and allowed_file(photo.filename):
                filename = secure_filename(f"{job_id}_{uuid.uuid4().hex[:6]}_{photo.filename}")
                photo.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                db.execute("INSERT INTO job_photos (job_id, photo_path) VALUES (?,?)",
                           (job_id, filename))

        log_action(db, job_id, 'Job Created & Sent for Inspection', session['user_id'],
                   f"Customer: {data['customer_name']}")

        # Send WhatsApp notification to customer
        job_data = dict(data)
        send_job_created_notification(job_data)

        # Notify the assigned technician
        if tech_id:
            tech_name = db.execute("SELECT name FROM users WHERE id=?", (tech_id,)).fetchone()
            tech_label = tech_name['name'] if tech_name else 'Technician'
            send_notification(
                db, tech_id, job_id,
                f"📦 New job {job_id} assigned to you for inspection. "
                f"Customer: {data['customer_name']} | Item: {data['item_type']}"
            )

        db.commit()
        flash(f'Job {job_id} created and sent for inspection! WhatsApp notification sent to customer.', 'success')
        return redirect(url_for('admin_job_detail', job_id=job_id))

    db = get_db()
    technicians = db.execute("SELECT * FROM users WHERE role='technician'").fetchall()
    return render_template('admin/new_job.html', technicians=technicians)

@app.route('/admin/jobs/<job_id>')
@admin_required
def admin_job_detail(job_id):
    db = get_db()
    job = db.execute("""
        SELECT j.*, u.name as tech_name FROM jobs j
        LEFT JOIN users u ON j.assigned_tech_id = u.id
        WHERE j.job_id=?
    """, (job_id,)).fetchone()
    if not job:
        flash('Job not found', 'error')
        return redirect(url_for('admin_jobs'))
    photos = db.execute("SELECT * FROM job_photos WHERE job_id=?", (job_id,)).fetchall()
    logs = db.execute("""
        SELECT l.*, u.name as user_name FROM job_logs l
        LEFT JOIN users u ON l.performed_by = u.id
        WHERE l.job_id=? ORDER BY l.created_at DESC
    """, (job_id,)).fetchall()
    technicians = db.execute("SELECT * FROM users WHERE role='technician'").fetchall()
    return render_template('admin/job_detail.html', job=job, photos=photos,
                           logs=logs, technicians=technicians,
                           status_flow=STATUS_FLOW, status_order=STATUS_ORDER)

@app.route('/admin/jobs/<job_id>/update', methods=['POST'])
@admin_required
def admin_update_job(job_id):
    db = get_db()
    action = request.form.get('action')
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    tech_id = get_assigned_tech_id(db, job_id)

    # ── Assign technician ────────────────────────────────────────────────────
    if action == 'assign_tech':
        new_tech_id = request.form.get('tech_id')
        db.execute("UPDATE jobs SET assigned_tech_id=?, updated_at=? WHERE job_id=?",
                   (new_tech_id, now, job_id))
        tech = db.execute("SELECT name FROM users WHERE id=?", (new_tech_id,)).fetchone()
        log_action(db, job_id, f'Assigned to {tech["name"]}', session['user_id'])
        if new_tech_id:
            send_notification(
                db, new_tech_id, job_id,
                f"📋 Job {job_id} has been assigned to you for inspection."
            )
        db.commit()
        flash('Technician assigned!', 'success')
        return redirect(url_for('admin_job_detail', job_id=job_id))

    # ── Approve estimate → sent_for_repair ──────────────────────────────────
    elif action == 'approve_estimate':
        db.execute("""
            UPDATE jobs SET status='estimate_approved', estimate_approved_at=?, updated_at=?
            WHERE job_id=?
        """, (now, now, job_id))
        log_action(db, job_id, 'Estimate Approved by Admin', session['user_id'])
        
        # Send WhatsApp confirmation to customer
        job = db.execute("SELECT * FROM jobs WHERE job_id=?", (job_id,)).fetchone()
        send_estimate_approved_confirmation(dict(job))
        
        # Immediately move to sent_for_repair
        db.execute("UPDATE jobs SET status='sent_for_repair', updated_at=? WHERE job_id=?",
                   (now, job_id))
        log_action(db, job_id, 'Status → sent_for_repair', session['user_id'])
        
        # Notify technician
        if tech_id:
            send_notification(
                db, tech_id, job_id,
                f"✅ Estimate approved for job {job_id}. Please proceed with the repair."
            )
        db.commit()
        flash('Estimate approved — job sent for repair and customer notified!', 'success')
        return redirect(url_for('admin_job_detail', job_id=job_id))

    # ── Reject estimate ──────────────────────────────────────────────────────
    elif action == 'reject_estimate':
        reason = request.form.get('rejection_reason', '')
        db.execute("""
            UPDATE jobs SET status='estimate_rejected', notes=?, updated_at=?
            WHERE job_id=?
        """, (reason, now, job_id))
        log_action(db, job_id, 'Estimate Rejected by Admin', session['user_id'], reason)
        
        # Send WhatsApp rejection confirmation
        job = db.execute("SELECT * FROM jobs WHERE job_id=?", (job_id,)).fetchone()
        send_estimate_rejected_confirmation(dict(job))
        
        if tech_id:
            send_notification(
                db, tech_id, job_id,
                f"🚫 Estimate for job {job_id} was rejected by admin. Reason: {reason}"
            )
        db.commit()
        flash('Estimate rejected and customer notified.', 'success')
        return redirect(url_for('admin_job_detail', job_id=job_id))

    # ── Upload invoice + select payment method + send (unified) ─────────────
    elif action == 'upload_invoice_and_send':
        invoice_file = request.files.get('invoice_file')
        invoice_number = request.form.get('invoice_number', '').strip()
        invoice_total_amount = float(request.form.get('invoice_total_amount', 0) or 0)
        payment_method = request.form.get('payment_method', 'razorpay')
        other_notes = request.form.get('other_payment_notes', '').strip()

        if not invoice_file or not allowed_file(invoice_file.filename):
            flash('Please upload a valid invoice file (PDF, PNG, JPG).', 'error')
            db.commit()
            return redirect(url_for('admin_job_detail', job_id=job_id))

        if invoice_total_amount <= 0 and payment_method != 'free_of_charge':
            flash('Invoice amount must be greater than 0.', 'error')
            db.commit()
            return redirect(url_for('admin_job_detail', job_id=job_id))

        # ── Save invoice file ──────────────────────────────────────────────────
        filename = secure_filename(f"INV_{job_id}_{uuid.uuid4().hex[:8]}_{invoice_file.filename}")
        invoice_file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        invoice_url = build_public_url(f'/static/uploads/{filename}')

        db.execute("""
            UPDATE jobs SET
                invoice_path=?, invoice_number=?,
                invoice_total_amount=?, total_amount=?,
                payment_method=?,
                status='invoice_uploaded', updated_at=?
            WHERE job_id=?
        """, (filename, invoice_number, invoice_total_amount,
              invoice_total_amount, payment_method, now, job_id))

        log_action(db, job_id,
                   f'Invoice Uploaded & Sent: {invoice_number or job_id} (₹{invoice_total_amount:.2f}) — {payment_method}',
                   session['user_id'])

        # Re-fetch job so payment_token etc. are available after update
        job = db.execute("SELECT * FROM jobs WHERE job_id=?", (job_id,)).fetchone()
        job_dict = dict(job)

        # ── Handle payment method ──────────────────────────────────────────────
        if payment_method == 'razorpay':
            amount_paise = int(invoice_total_amount * 100)
            try:
                order = razorpay_client.order.create({
                    'amount': amount_paise, 'currency': 'INR', 'receipt': job_id,
                    'notes': {'job_id': job_id, 'customer': job['customer_name'],
                              'invoice': invoice_number or job_id}
                })
                token = uuid.uuid4().hex
                pay_link = build_public_url(f'/pay/{job_id}/{token}')
                db.execute("""
                    UPDATE jobs SET razorpay_order_id=?, payment_token=?, payment_link=?, updated_at=?
                    WHERE job_id=?
                """, (order['id'], token, pay_link, now, job_id))
                job_dict['payment_token'] = token  # inject token for template params
                send_invoice_ready_notification(
                    job_dict, invoice_total_amount,
                    job['parts_cost'] or 0, job['labour_cost'] or 0,
                    invoice_url
                )
                flash(f'✅ Invoice sent with Razorpay payment link via WhatsApp!', 'success')
            except Exception as e:
                flash(f'Invoice saved but Razorpay error: {str(e)}', 'error')

        elif payment_method == 'cash_cheque':
            # Step 1: Send non-razorpay invoice message
            send_invoice_ready_notification_non_razorpay(
                job_dict, invoice_total_amount,
                job['parts_cost'] or 0, job['labour_cost'] or 0,
                'Cash / Cheque',
                invoice_url
            )
            log_action(db, job_id, 'Payment Method: Cash / Cheque — Invoice Sent', session['user_id'])
            # Step 2: Mark as paid and send payment received message
            db.execute("""
                UPDATE jobs SET payment_status='paid', payment_received_at=?,
                status='payment_received', updated_at=? WHERE job_id=?
            """, (now, now, job_id))
            log_action(db, job_id, 'Auto-Marked Paid (Cash / Cheque)', session['user_id'])
            send_payment_received_confirmation(job_dict, invoice_total_amount, 'CASH_CHEQUE')
            flash('✅ Invoice sent and payment marked received via WhatsApp (Cash / Cheque).', 'success')

        elif payment_method == 'pay_later':
            # Step 1: Send non-razorpay invoice message
            send_invoice_ready_notification_non_razorpay(
                job_dict, invoice_total_amount,
                job['parts_cost'] or 0, job['labour_cost'] or 0,
                'Pay Later',
                invoice_url
            )
            log_action(db, job_id, 'Payment Method: Pay Later — Invoice Sent', session['user_id'],
                       f'Amount ₹{invoice_total_amount:.2f} deferred')
            # Step 2: Mark as paid, move to dispatched (no payment received message)
            db.execute("""
                UPDATE jobs SET payment_status='paid', payment_received_at=?,
                status='payment_received', updated_at=? WHERE job_id=?
            """, (now, now, job_id))
            log_action(db, job_id, 'Auto-Marked Paid & Dispatched (Pay Later)', session['user_id'])
            flash('✅ Invoice sent via WhatsApp. Job moved to Dispatched (Pay Later).', 'success')

        elif payment_method == 'free_of_charge':
            # Mark as paid, move to dispatched (no payment received message)
            db.execute("""
                UPDATE jobs SET payment_status='paid', payment_received_at=?,
                status='payment_received', updated_at=? WHERE job_id=?
            """, (now, now, job_id))
            log_action(db, job_id, 'Free of Charge — Payment Waived & Dispatched', session['user_id'])
            flash('✅ Marked Free of Charge. Job moved to Dispatched.', 'success')

        elif payment_method == 'other':
            # Step 1: Send non-razorpay invoice message
            send_invoice_ready_notification_non_razorpay(
                job_dict, invoice_total_amount,
                job['parts_cost'] or 0, job['labour_cost'] or 0,
                other_notes or 'Other',
                invoice_url
            )
            log_action(db, job_id, 'Payment Method: Other — Invoice Sent', session['user_id'],
                       other_notes or 'No notes')
            # Step 2: Mark as paid, move to dispatched (no payment received message)
            db.execute("""
                UPDATE jobs SET payment_status='paid', payment_received_at=?,
                status='payment_received', updated_at=? WHERE job_id=?
            """, (now, now, job_id))
            log_action(db, job_id, 'Auto-Marked Paid & Dispatched (Other)', session['user_id'])
            flash(f'✅ Invoice sent via WhatsApp. Job moved to Dispatched ({other_notes or "Other"}).', 'success')

        db.commit()
        return redirect(url_for('admin_job_detail', job_id=job_id))

    # ── Manual payment received ──────────────────────────────────────────────
    elif action == 'payment_received':
        db.execute("""
            UPDATE jobs SET payment_status='paid', payment_received_at=?,
            status='payment_received', updated_at=? WHERE job_id=?
        """, (now, now, job_id))
        log_action(db, job_id, 'Payment Received (Manual)', session['user_id'])
        
        # Send WhatsApp payment confirmation
        job = db.execute("SELECT * FROM jobs WHERE job_id=?", (job_id,)).fetchone()
        payment_method = job.get('payment_method', 'manual').upper()
        send_payment_received_confirmation(
            dict(job),
            job['total_amount'] or 0,
            f'MANUAL-{payment_method}'   # e.g. "MANUAL-CASH_CHEQUE"
        )
        
        db.commit()
        flash('Payment marked as received and customer notified!', 'success')
        return redirect(url_for('admin_job_detail', job_id=job_id))

    # ── Dispatch ─────────────────────────────────────────────────────────────
    elif action == 'dispatch':
        tracking = request.form.get('tracking_number')
        courier_name = request.form.get('courier_name')
        dispatch_date = request.form.get('dispatch_date')
        expected = request.form.get('expected_delivery')

        # ✅ Save the courier receipt file
        receipt_file = request.files.get('courier_receipt')
        receipt_path = None
        receipt_public_url = None

        if receipt_file and receipt_file.filename:
            ext = secure_filename(receipt_file.filename).rsplit('.', 1)[-1].lower()
            filename = f"lr_{job_id}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
            save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            receipt_file.save(save_path)
            receipt_path = filename
            receipt_public_url = build_public_url(f'/static/uploads/{filename}')

        db.execute("""
            UPDATE jobs SET status='dispatched', tracking_number=?, courier_name=?,
            dispatch_date=?, expected_delivery=?, courier_receipt_path=?, updated_at=?
            WHERE job_id=?
        """, (tracking, courier_name, dispatch_date, expected, receipt_path, now, job_id))
        log_action(db, job_id, 'Dispatched', session['user_id'],
                   f'Courier: {courier_name}, Tracking: {tracking}')

        db.commit()  # ✅ commit before fetching

        # ✅ Send WhatsApp with receipt as media
        job = db.execute("SELECT * FROM jobs WHERE job_id=?", (job_id,)).fetchone()
        send_dispatched_notification(dict(job), courier_name, tracking, expected, receipt_public_url)

        flash('Job dispatched and customer notified via WhatsApp!', 'success')
        return redirect(url_for('admin_job_detail', job_id=job_id))

    # ── Close job ────────────────────────────────────────────────────────────
    elif action == 'close':
        db.execute("UPDATE jobs SET status='closed', updated_at=? WHERE job_id=?", (now, job_id))
        log_action(db, job_id, 'Job Closed', session['user_id'])
        db.commit()
        flash('Job closed.', 'success')
        return redirect(url_for('admin_job_detail', job_id=job_id))

    db.commit()
    flash('Job updated successfully!', 'success')
    return redirect(url_for('admin_job_detail', job_id=job_id))

@app.route('/admin/jobs/<job_id>/delete', methods=['POST'])
@admin_required
def admin_delete_job(job_id):
    db = get_db()
    db.execute("DELETE FROM job_photos WHERE job_id=?", (job_id,))
    db.execute("DELETE FROM job_logs WHERE job_id=?", (job_id,))
    db.execute("DELETE FROM notifications WHERE job_id=?", (job_id,))
    db.execute("DELETE FROM jobs WHERE job_id=?", (job_id,))
    db.commit()
    flash(f'Job {job_id} deleted.', 'success')
    return redirect(url_for('admin_jobs'))

# ─── USER MANAGEMENT (Admin only) ────────────────────────────────────────────
ALL_PERMISSIONS = {
    'view_all_jobs':   'View All Jobs',
    'create_jobs':     'Create Jobs',
    'edit_jobs':       'Edit Jobs',
    'delete_jobs':     'Delete Jobs',
    'upload_invoice':  'Upload Invoice',
    'manage_payments': 'Manage Payments',
    'dispatch_jobs':   'Dispatch Jobs',
    'view_reports':    'View Reports',
    'manage_users':    'Manage Users',
    'export_data':     'Export Data',
}

@app.route('/admin/users')
@admin_required
def admin_users():
    db = get_db()
    rows = db.execute("""
        SELECT u.*, COUNT(j.id) AS job_count
        FROM users u
        LEFT JOIN jobs j ON j.assigned_tech_id = u.id
        GROUP BY u.id
        ORDER BY u.role, u.name
    """).fetchall()
    users = []
    for r in rows:
        d = dict(r)
        try:
            d['perms'] = json.loads(r['permissions'] or '{}')
        except Exception:
            d['perms'] = {}
        users.append(d)
    return render_template('admin/users.html', users=users, all_permissions=ALL_PERMISSIONS)

@app.route('/admin/users/add', methods=['POST'])
@admin_required
def admin_add_user():
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    name     = request.form.get('name', '').strip()
    role     = request.form.get('role', 'technician')

    if not username or not password or not name:
        flash('Username, password and name are required.', 'error')
        return redirect(url_for('admin_users'))

    db = get_db()
    if db.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone():
        flash(f'Username "{username}" is already taken.', 'error')
        return redirect(url_for('admin_users'))

    perms = {k: True for k in ALL_PERMISSIONS if request.form.get(k)}
    db.execute(
        "INSERT INTO users (username, password, role, name, permissions) VALUES (?,?,?,?,?)",
        (username, generate_password_hash(password), role, name, json.dumps(perms))
    )
    db.commit()
    flash(f'User "{name}" created successfully!', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/delete/<int:user_id>', methods=['POST'])
@admin_required
def admin_delete_user(user_id):
    if user_id == session['user_id']:
        flash("You can't delete your own account.", 'error')
        return redirect(url_for('admin_users'))
    db = get_db()
    user = db.execute("SELECT name FROM users WHERE id=?", (user_id,)).fetchone()
    if user:
        db.execute("DELETE FROM users WHERE id=?", (user_id,))
        db.commit()
        flash(f'User "{user["name"]}" deleted.', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/permissions/<int:user_id>', methods=['POST'])
@admin_required
def admin_save_permissions(user_id):
    perms = {k: True for k in ALL_PERMISSIONS if request.form.get(k)}
    db = get_db()
    db.execute("UPDATE users SET permissions=? WHERE id=?", (json.dumps(perms), user_id))
    db.commit()
    flash('Permissions saved successfully.', 'success')
    return redirect(url_for('admin_users'))

# ─── BACKUP & RESTORE (Admin only) ───────────────────────────────────────────
import zipfile, shutil, tempfile

@app.route('/admin/backup')
@admin_required
def admin_backup_page():
    """Admin-only backup & restore page."""
    return render_template('admin/backup.html')

@app.route('/admin/backup/download')
@admin_required
def admin_backup_download():
    """Create and stream a ZIP backup of the database + uploads folder."""
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    zip_name = f'maktronics_backup_{ts}.zip'
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Database
        if os.path.exists(DB_PATH):
            zf.write(DB_PATH, 'maktronics.db')
        # Uploads folder
        upload_dir = app.config['UPLOAD_FOLDER']
        if os.path.exists(upload_dir):
            for root, dirs, files in os.walk(upload_dir):
                for file in files:
                    abs_path = os.path.join(root, file)
                    arc_name = os.path.relpath(abs_path, start=os.path.dirname(upload_dir))
                    zf.write(abs_path, arc_name)
    buf.seek(0)
    return send_file(buf, mimetype='application/zip',
                     as_attachment=True, download_name=zip_name)

@app.route('/admin/backup/restore', methods=['POST'])
@admin_required
def admin_backup_restore():
    """Restore database from an uploaded ZIP backup."""
    f = request.files.get('backup_file')
    if not f or not f.filename.endswith('.zip'):
        flash('Please upload a valid .zip backup file.', 'error')
        return redirect(url_for('admin_backup_page'))

    try:
        buf = io.BytesIO(f.read())
        with zipfile.ZipFile(buf, 'r') as zf:
            names = zf.namelist()
            if 'maktronics.db' not in names:
                flash('Invalid backup: maktronics.db not found inside ZIP.', 'error')
                return redirect(url_for('admin_backup_page'))

            # Backup current DB before overwriting
            if os.path.exists(DB_PATH):
                ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
                shutil.copy2(DB_PATH, f'{DB_PATH}.pre_restore_{ts}.bak')

            # Restore DB
            with zf.open('maktronics.db') as src:
                with open(DB_PATH, 'wb') as dst:
                    shutil.copyfileobj(src, dst)

            # Restore uploads (optional — only if they're in the ZIP)
            upload_dir = app.config['UPLOAD_FOLDER']
            for name in names:
                if name.startswith('static/uploads/') and not name.endswith('/'):
                    dest = os.path.join(os.path.dirname(upload_dir),
                                        *name.split('/')[1:])  # strip 'static/'
                    os.makedirs(os.path.dirname(dest), exist_ok=True)
                    with zf.open(name) as src, open(dest, 'wb') as dst:
                        shutil.copyfileobj(src, dst)

        flash('✅ Backup restored successfully! The database has been replaced.', 'success')
    except Exception as e:
        flash(f'Restore failed: {str(e)}', 'error')
    return redirect(url_for('admin_backup_page'))

# ─── TECHNICIAN ROUTES ────────────────────────────────────────────────────────
@app.route('/tech')
@login_required
def tech_dashboard():
    db = get_db()
    tech_id = session['user_id']
    stats = {
        'total':        db.execute("SELECT COUNT(*) FROM jobs WHERE assigned_tech_id=?", (tech_id,)).fetchone()[0],
        'sent_for_repair': db.execute("SELECT COUNT(*) FROM jobs WHERE assigned_tech_id=? AND status='sent_for_repair'", (tech_id,)).fetchone()[0],
        'repair_done':  db.execute("SELECT COUNT(*) FROM jobs WHERE assigned_tech_id=? AND status='repair_done'", (tech_id,)).fetchone()[0],
        'closed':       db.execute("SELECT COUNT(*) FROM jobs WHERE assigned_tech_id=? AND status='closed'", (tech_id,)).fetchone()[0],
    }
    my_jobs = db.execute("""
        SELECT * FROM jobs WHERE assigned_tech_id=? ORDER BY received_at DESC LIMIT 10
    """, (tech_id,)).fetchall()
    return render_template('technician/dashboard.html', stats=stats, jobs=my_jobs,
                           status_flow=STATUS_FLOW)

@app.route('/tech/jobs')
@login_required
def tech_jobs():
    db = get_db()
    tech_id = session['user_id']
    status_filter = request.args.get('status', '')
    search = request.args.get('search', '')
    query = "SELECT * FROM jobs WHERE assigned_tech_id=?"
    params = [tech_id]
    if status_filter:
        query += " AND status=?"
        params.append(status_filter)
    if search:
        query += " AND (job_id LIKE ? OR barcode LIKE ? OR customer_name LIKE ?)"
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])
    query += " ORDER BY received_at DESC"
    jobs = db.execute(query, params).fetchall()
    return render_template('technician/jobs.html', jobs=jobs, status_flow=STATUS_FLOW,
                           status_filter=status_filter, search=search)

@app.route('/tech/jobs/<job_id>')
@login_required
def tech_job_detail(job_id):
    db = get_db()
    tech_id = session['user_id']
    if session['role'] == 'admin':
        job = db.execute(
            "SELECT j.*, u.name as tech_name FROM jobs j LEFT JOIN users u ON j.assigned_tech_id=u.id WHERE j.job_id=?",
            (job_id,)
        ).fetchone()
    else:
        job = db.execute(
            "SELECT * FROM jobs WHERE job_id=? AND assigned_tech_id=?", (job_id, tech_id)
        ).fetchone()
    if not job:
        flash('Job not found or not assigned to you', 'error')
        return redirect(url_for('tech_jobs'))
    photos = db.execute("SELECT * FROM job_photos WHERE job_id=?", (job_id,)).fetchall()
    logs = db.execute("""
        SELECT l.*, u.name as user_name FROM job_logs l
        LEFT JOIN users u ON l.performed_by=u.id
        WHERE l.job_id=? ORDER BY l.created_at DESC
    """, (job_id,)).fetchall()
    return render_template('technician/job_detail.html', job=job, photos=photos,
                           logs=logs, status_flow=STATUS_FLOW, status_order=STATUS_ORDER)

@app.route('/tech/jobs/<job_id>/update', methods=['POST'])
@login_required
def tech_update_job(job_id):
    db = get_db()
    tech_id = session['user_id']
    action = request.form.get('action')
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Ensure technician owns this job
    if session['role'] != 'admin':
        job = db.execute(
            "SELECT * FROM jobs WHERE job_id=? AND assigned_tech_id=?", (job_id, tech_id)
        ).fetchone()
        if not job:
            flash('Access denied', 'error')
            return redirect(url_for('tech_jobs'))

    job = db.execute("SELECT * FROM jobs WHERE job_id=?", (job_id,)).fetchone()

    # ── Mark not repairable ──────────────────────────────────────────────────
    if action == 'not_repairable':
        reason = request.form.get('not_repairable_reason', '')
        findings = request.form.get('inspection_findings', '')
        db.execute("""
            UPDATE jobs SET status='not_repairable', inspection_findings=?,
            not_repairable_reason=?, updated_at=? WHERE job_id=?
        """, (findings, reason, now, job_id))
        log_action(db, job_id, 'Marked Not Repairable', tech_id, reason)
        
        # Send WhatsApp notification to customer
        send_not_repairable_notification(dict(job), reason)
        
        notify_all_admins(
            db, job_id,
            f"❌ Job {job_id} marked NOT REPAIRABLE by {session['name']}. "
            f"Reason: {reason}"
        )

    # ── Send estimate ────────────────────────────────────────────────────────
    elif action == 'send_estimate':
        estimate_amount = float(request.form.get('estimate_amount', 0))
        parts_cost = float(request.form.get('parts_cost', 0))
        labour_cost = float(request.form.get('labour_cost', 0))
        estimate_notes = request.form.get('estimate_notes', '')
        findings = request.form.get('inspection_findings', '')
        db.execute("""
            UPDATE jobs SET status='estimate_sent', inspection_findings=?,
            estimate_amount=?, parts_cost=?, labour_cost=?,
            estimate_notes=?, estimate_sent_at=?, updated_at=?
            WHERE job_id=?
        """, (findings, estimate_amount, parts_cost, labour_cost, estimate_notes, now, now, job_id))
        log_action(db, job_id, f'Estimate Sent: ₹{estimate_amount}', tech_id, estimate_notes)

        # Re-fetch job AFTER update so dict has fresh data
        job = db.execute("SELECT * FROM jobs WHERE job_id=?", (job_id,)).fetchone()

        # Send WhatsApp estimate for customer approval
        result = send_estimate_notification(
            dict(job),
            float(job['estimate_amount'] or 0),
            float(job['parts_cost'] or 0),
            float(job['labour_cost'] or 0)
        )
        print(f"[WhatsApp Estimate] result: {result}")

        notify_all_admins(
            db, job_id,
            f"📄 Technician {session['name']} sent estimate ₹{estimate_amount:.2f} "
            f"for job {job_id}. Customer will approve via WhatsApp."
        )

    # ── Repair done ──────────────────────────────────────────────────────────
    elif action == 'repair_done':
        findings = request.form.get('repair_findings', '')
        parts = float(request.form.get('parts_cost', 0))
        labour = float(request.form.get('labour_cost', 0))
        total = parts + labour
        db.execute("""
            UPDATE jobs SET status='repair_done', repair_findings=?,
            parts_cost=?, labour_cost=?, total_amount=?, updated_at=?
            WHERE job_id=?
        """, (findings, parts, labour, total, now, job_id))
        log_action(db, job_id, 'Repair Done', tech_id,
                   f'Findings: {findings} | Total: ₹{total}')

        # Re-fetch job after update for accurate data
        job = db.execute("SELECT * FROM jobs WHERE job_id=?", (job_id,)).fetchone()

        # Notify accounts department to generate invoice
        accounts_result = send_accounts_department_notification(dict(job), total, parts, labour, session.get('name', 'Technician'))
        if accounts_result.get('success'):
            print(f"[Accounts] Invoice request sent to accounts for job {job_id}")
        else:
            print(f"[Accounts] Failed to notify accounts for job {job_id}: {accounts_result.get('error')}")

        notify_all_admins(
            db, job_id,
            f"✅ Repair completed for job {job_id} by {session['name']}. "
            f"Total: ₹{total:.2f}. Accounts department notified to generate invoice."
        )

    # ── Add photo ────────────────────────────────────────────────────────────
    elif action == 'add_photo':
        photos = request.files.getlist('photos')
        for photo in photos[:3]:
            if photo and allowed_file(photo.filename):
                filename = secure_filename(f"{job_id}_{uuid.uuid4().hex[:6]}_{photo.filename}")
                photo.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                db.execute("INSERT INTO job_photos (job_id, photo_path) VALUES (?,?)",
                           (job_id, filename))
        log_action(db, job_id, 'Photos Added', tech_id)

    # ── Progress note ────────────────────────────────────────────────────────
    elif action == 'update_progress':
        notes = request.form.get('notes')
        db.execute("UPDATE jobs SET notes=?, updated_at=? WHERE job_id=?", (notes, now, job_id))
        log_action(db, job_id, 'Progress Note Added', tech_id, notes)

    db.commit()
    flash('Job updated!', 'success')
    return redirect(url_for('tech_job_detail', job_id=job_id))

# ─── BARCODE SCANNER ──────────────────────────────────────────────────────────
@app.route('/scanner')
@login_required
def scanner():
    return render_template('shared/scanner.html')

@app.route('/api/scan', methods=['GET'])
@login_required
def api_scan():
    code = request.args.get('code', '').strip()
    if not code:
        return jsonify({'found': False, 'error': 'No code provided'}), 400
    db = get_db()
    job = db.execute(
        "SELECT job_id FROM jobs WHERE job_id=? OR barcode=?", (code, code)
    ).fetchone()
    if not job:
        return jsonify({'found': False, 'error': f'No job found for: {code}'}), 404
    role = session.get('role')
    if role in ('admin', 'manager'):
        url = url_for('admin_job_detail', job_id=job['job_id'])
    else:
        url = url_for('tech_job_detail', job_id=job['job_id'])
    return jsonify({'found': True, 'job_id': job['job_id'], 'redirect': url})

# ─── CUSTOMER PAYMENT PAGE ────────────────────────────────────────────────────
@app.route('/pay/<job_id>/<token>')
def customer_pay(job_id, token):
    db = get_db()
    job = db.execute(
        "SELECT * FROM jobs WHERE job_id=? AND payment_token=?", (job_id, token)
    ).fetchone()
    if not job:
        return "Invalid or expired payment link.", 404
    if job['payment_status'] == 'paid':
        return render_template('shared/payment_success.html', job=job, already_paid=True)
    amount = float(job['invoice_total_amount'] or job['total_amount'] or 0)
    return render_template('shared/customer_payment.html', job=job, amount=amount,
                           razorpay_key=RAZORPAY_KEY_ID)

@app.route('/pay/<job_id>/verify', methods=['POST'])
def verify_payment(job_id):
    db = get_db()
    data = request.get_json() or request.form.to_dict()
    razorpay_order_id   = data.get('razorpay_order_id', '')
    razorpay_payment_id = data.get('razorpay_payment_id', '')
    razorpay_signature  = data.get('razorpay_signature', '')

    body = f"{razorpay_order_id}|{razorpay_payment_id}".encode()
    expected_sig = hmac.new(RAZORPAY_KEY_SECRET.encode(), body, hashlib.sha256).hexdigest()
    if not hmac.compare_digest(expected_sig, razorpay_signature):
        return jsonify({'status': 'error', 'message': 'Signature mismatch'}), 400

    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    db.execute("""
        UPDATE jobs SET payment_status='paid', payment_received_at=?,
        razorpay_payment_id=?, status='payment_received', updated_at=?
        WHERE job_id=?
    """, (now, razorpay_payment_id, now, job_id))
    db.execute(
        "INSERT INTO job_logs (job_id, action, performed_by, details) VALUES (?,?,?,?)",
        (job_id, 'Payment Received via Razorpay', None,
         f'Payment ID: {razorpay_payment_id}, Order: {razorpay_order_id}')
    )
    
    # Send WhatsApp payment confirmation
    job = db.execute("SELECT * FROM jobs WHERE job_id=?", (job_id,)).fetchone()
    send_payment_received_confirmation(dict(job), job['total_amount'] or 0, razorpay_payment_id)
    
    db.commit()
    return jsonify({'status': 'success', 'redirect': url_for('payment_success', job_id=job_id)})

@app.route('/pay/<job_id>/success')
def payment_success(job_id):
    db = get_db()
    job = db.execute("SELECT * FROM jobs WHERE job_id=?", (job_id,)).fetchone()
    if not job:
        return "Job not found.", 404
    return render_template('shared/payment_success.html', job=job, already_paid=False)

@app.route('/razorpay/webhook', methods=['POST'])
def razorpay_webhook():
    webhook_secret = os.environ.get('RAZORPAY_WEBHOOK_SECRET', '')
    payload = request.get_data()
    received_sig = request.headers.get('X-Razorpay-Signature', '')
    if webhook_secret:
        expected = hmac.new(webhook_secret.encode(), payload, hashlib.sha256).hexdigest()
        if not hmac.compare_digest(expected, received_sig):
            return 'Invalid signature', 400
    event = request.json
    if event and event.get('event') == 'payment.captured':
        payment = event['payload']['payment']['entity']
        order_id  = payment.get('order_id')
        payment_id = payment.get('id')
        db = get_db()
        job = db.execute("SELECT * FROM jobs WHERE razorpay_order_id=?", (order_id,)).fetchone()
        if job and job['payment_status'] != 'paid':
            now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            db.execute("""
                UPDATE jobs SET payment_status='paid', payment_received_at=?,
                razorpay_payment_id=?, status='payment_received', updated_at=?
                WHERE razorpay_order_id=?
            """, (now, payment_id, now, order_id))
            db.execute(
                "INSERT INTO job_logs (job_id, action, performed_by, details) VALUES (?,?,?,?)",
                (job['job_id'], 'Payment Confirmed via Webhook', None, f'Payment ID: {payment_id}')
            )
            
            # Send WhatsApp payment confirmation
            send_payment_received_confirmation(dict(job), job['total_amount'] or 0, payment_id)
            
            db.commit()
    return 'OK', 200

# ─── WHATSAPP SETTINGS PAGE (Optional) ────────────────────────────────────────
@app.route('/admin/whatsapp')
@admin_required
def whatsapp_settings():
    aisensy_configured = bool(AISENSY_API_KEY)
    return render_template('admin/whatsapp_settings.html', aisensy_configured=aisensy_configured)

# ─── TEST WHATSAPP ENDPOINT (Development only) ───────────────────────────────
@app.route('/admin/test-whatsapp/<job_id>')
@admin_required
def test_whatsapp(job_id):
    """Test endpoint to manually send WhatsApp messages"""
    db = get_db()
    job = db.execute("SELECT * FROM jobs WHERE job_id=?", (job_id,)).fetchone()
    if not job:
        flash('Job not found', 'error')
        return redirect(url_for('admin_dashboard'))
    
    result = send_job_created_notification(dict(job))
    if result.get('success'):
        flash(f'Test WhatsApp sent to {job["customer_phone"]}', 'success')
    else:
        flash(f'WhatsApp test failed: {result.get("error")}', 'error')
    
    return redirect(url_for('admin_job_detail', job_id=job_id))

def migrate_db():
    """Add any missing columns to existing databases (safe to run repeatedly)."""
    with get_db() as db:
        # Get existing columns in jobs table
        existing_cols = {row[1] for row in db.execute("PRAGMA table_info(jobs)").fetchall()}

        migrations = [
            ("payment_method",          "ALTER TABLE jobs ADD COLUMN payment_method TEXT DEFAULT 'razorpay'"),
            ("invoice_generate_date",   "ALTER TABLE jobs ADD COLUMN invoice_generate_date TIMESTAMP"),
            ("invoice_total_amount",    "ALTER TABLE jobs ADD COLUMN invoice_total_amount REAL DEFAULT 0"),
            ("razorpay_order_id",       "ALTER TABLE jobs ADD COLUMN razorpay_order_id TEXT"),
            ("razorpay_payment_id",     "ALTER TABLE jobs ADD COLUMN razorpay_payment_id TEXT"),
            ("payment_link",            "ALTER TABLE jobs ADD COLUMN payment_link TEXT"),
            ("payment_token",           "ALTER TABLE jobs ADD COLUMN payment_token TEXT"),
            ("whatsapp_message_id",     "ALTER TABLE jobs ADD COLUMN whatsapp_message_id TEXT"),
            ("estimate_amount",         "ALTER TABLE jobs ADD COLUMN estimate_amount REAL DEFAULT 0"),
            ("estimate_notes",          "ALTER TABLE jobs ADD COLUMN estimate_notes TEXT"),
            ("estimate_sent_at",        "ALTER TABLE jobs ADD COLUMN estimate_sent_at TIMESTAMP"),
            ("estimate_approved_at",    "ALTER TABLE jobs ADD COLUMN estimate_approved_at TIMESTAMP"),
            ("not_repairable_reason",   "ALTER TABLE jobs ADD COLUMN not_repairable_reason TEXT"),
            ("sent_back_to_customer_at","ALTER TABLE jobs ADD COLUMN sent_back_to_customer_at TIMESTAMP"),
            ("inspection_findings",     "ALTER TABLE jobs ADD COLUMN inspection_findings TEXT"),
            ("courier_name",            "ALTER TABLE jobs ADD COLUMN courier_name TEXT"),
            ("courier_receipt_path",    "ALTER TABLE jobs ADD COLUMN courier_receipt_path TEXT"),
        ]

        for col_name, sql in migrations:
            if col_name not in existing_cols:
                try:
                    db.execute(sql)
                    print(f"[migrate_db] Added column: {col_name}")
                except Exception as e:
                    print(f"[migrate_db] Skipped {col_name}: {e}")

        db.commit()

init_db()
migrate_db()

if __name__ == '__main__':
    app.run(debug=True, port=5015)
